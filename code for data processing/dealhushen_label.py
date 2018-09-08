#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Created on Sun Jul 29 16:17:19 2018

@author: luffy
"""

#import pandas as pd
#import numpy as np
#from openpyxl import load_workbook
#excel_path = '/Users/luffy/Desktop/DL/data/hushen300/hushen300_data_deal_labelinitial.xlsx'  #hushen_deallabel.xlsx
#temp=pd.read_excel(excel_path, sheetname='pct_chg_train')
##chg_list=[]
#chg_label=[]
#for j in range(len(temp.columns)):
#    if j==0:
#        continue
#    else:
#        for i in range(int(len(temp)/30)):
#            temp_chg=np.mean(temp[temp.columns[j]][(30*i+30):(30*i+33)])
#            #chg_list.append(temp_chg)
#            if temp_chg < 0.363:  #0.328
#                chg_label.append(1)
#            elif temp_chg > 0.512: #0.5
#                chg_label.append(3)
#            else:
#                chg_label.append(2)
#temp3=pd.DataFrame(chg_label)
#def excelAddSheet(dataframe,excelWriter):
#    book = load_workbook(excelWriter.path)
#    excelWriter.book = book
#    dataframe.to_excel(excel_writer=excelWriter,sheet_name='pct_chg_train',index=None,columns=None)
#    excelWriter.close()
#excelPath='/Users/luffy/Desktop/DL/data/hushen300/hushen300_data_deal_label.xlsx'   #hushen_label.xlsx
#excelWriter=pd.ExcelWriter(excelPath,engine='openpyxl')
#excelAddSheet(temp3,excelWriter)


import pandas as pd
import numpy as np
from openpyxl import load_workbook
excel_path = '/Users/luffy/Desktop/DL/data/hushen300/hushen300_data_deal_labelinitial.xlsx'   #hushen_deallabel.xlsx
temp=pd.read_excel(excel_path, sheetname='pct_chg_test')
#chg_list=[]
chg_label=[]
for j in range(len(temp.columns)):
    if j==0:
        continue
    else:
        for i in range(int(len(temp)/30)):
            temp_chg=np.mean(temp[temp.columns[j]][(30*i+30):(30*i+33)])
            #chg_list.append(temp_chg)
            if temp_chg < 0.363:
                chg_label.append(1)
            elif temp_chg > 0.512:
                chg_label.append(3)
            else:
                chg_label.append(2)
temp3=pd.DataFrame(chg_label)
def excelAddSheet(dataframe,excelWriter):
    book = load_workbook(excelWriter.path)
    excelWriter.book = book
    dataframe.to_excel(excel_writer=excelWriter,sheet_name='pct_chg_test',index=None,columns=None)
    excelWriter.close()
excelPath='/Users/luffy/Desktop/DL/data/hushen300/hushen300_data_deal_label.xlsx'    #hushen_label.xlsx
excelWriter=pd.ExcelWriter(excelPath,engine='openpyxl')
excelAddSheet(temp3,excelWriter)