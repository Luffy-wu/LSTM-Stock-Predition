#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Created on Sun Jul 29 14:54:54 2018

@author: luffy
"""
import pandas as pd
import numpy as np
from openpyxl import load_workbook
excel_path = '/Users/luffy/Desktop/DL/data/hushen300/hushen300_data_deal_cut_train.xlsx' #hushen_cut_train.xlsx
namelist=['open','high','low','close','volume','amt','pct_chg','turn']
for j in range(len(namelist)):
    temp=pd.read_excel(excel_path, sheetname=namelist[j])
    temp_pd=pd.DataFrame(columns=[i for i in range(30)])
    temp2=temp.drop([2790,2791])
    temp2=temp2.drop(['Date'],axis=1)
    temp2=temp2.T
    temp3=np.array(temp2)
    temp3=temp3.reshape(-1,30)
    temp3=pd.DataFrame(temp3)
    def excelAddSheet(dataframe,excelWriter):
        book = load_workbook(excelWriter.path)
        excelWriter.book = book
        dataframe.to_excel(excel_writer=excelWriter,sheet_name=namelist[j],index=None,columns=None)
        excelWriter.close()
    excelPath='/Users/luffy/Desktop/DL/data/hushen300/hushen300_data_deal_sort_train.xlsx' #hushen_sort_train.xlsx
    excelWriter=pd.ExcelWriter(excelPath,engine='openpyxl')
    excelAddSheet(temp3,excelWriter)
    
#import pandas as pd
#import numpy as np
#from openpyxl import load_workbook
#excel_path = '/Users/luffy/Desktop/DL/data/hushen300/hushen300_data_deal_cut_test.xlsx'    #hushen_cut_test.xlsx
#namelist=['open','high','low','close','volume','amt','pct_chg','turn']
#for j in range(len(namelist)):
#    temp=pd.read_excel(excel_path, sheetname=namelist[j])
#    temp_pd=pd.DataFrame(columns=[i for i in range(30)])
#    drop_list=[i for i in range(240,264)]
#    temp2=temp.drop(drop_list)
#    temp2=temp2.drop(['Date'],axis=1)
#    temp2=temp2.T
#    temp3=np.array(temp2)
#    temp3=temp3.reshape(-1,30)
#    temp3=pd.DataFrame(temp3)
#    def excelAddSheet(dataframe,excelWriter):
#        book = load_workbook(excelWriter.path)
#        excelWriter.book = book
#        dataframe.to_excel(excel_writer=excelWriter,sheet_name=namelist[j],index=None,columns=None)
#        excelWriter.close()
#    excelPath='/Users/luffy/Desktop/DL/data/hushen300/hushen300_data_deal_sort_test.xlsx'   #hushen_sort_test.xlsx
#    excelWriter=pd.ExcelWriter(excelPath,engine='openpyxl')
#    excelAddSheet(temp3,excelWriter)