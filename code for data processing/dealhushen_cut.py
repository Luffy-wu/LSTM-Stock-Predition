#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Created on Sun Jul 29 14:38:29 2018

@author: luffy
"""
####revised on 2018.8.1

###train data
#import pandas as pd
#import numpy as np
#from openpyxl import load_workbook
#excel_path = '/Users/luffy/Desktop/DL/data/hushen300/hushen300_data_deal_standarize.xlsx'
#namelist=['open','high','low','close','volume','amt','pct_chg','turn']
#for j in range(len(namelist)):
#    temp = pd.read_excel(excel_path, sheetname=namelist[j])
#    temp2=temp[:2792]
#    def excelAddSheet(dataframe,excelWriter):
#        book = load_workbook(excelWriter.path)
#        excelWriter.book = book
#        dataframe.to_excel(excel_writer=excelWriter,sheet_name=namelist[j],index=None)
#        excelWriter.close()
#    excelPath='/Users/luffy/Desktop/DL/data/hushen300/hushen300_data_deal_cut_train.xlsx'
#    excelWriter=pd.ExcelWriter(excelPath,engine='openpyxl')
#    excelAddSheet(temp2,excelWriter)

##test data
import pandas as pd
import numpy as np
from openpyxl import load_workbook
excel_path = '/Users/luffy/Desktop/DL/data/hushen300/hushen300_data_deal_standarize.xlsx'
namelist=['open','high','low','close','volume','amt','pct_chg','turn']
for j in range(len(namelist)):
    temp = pd.read_excel(excel_path, sheetname=namelist[j])
    temp2=temp[2792:]
    def excelAddSheet(dataframe,excelWriter):
        book = load_workbook(excelWriter.path)
        excelWriter.book = book
        dataframe.to_excel(excel_writer=excelWriter,sheet_name=namelist[j],index=None)
        excelWriter.close()
    excelPath='/Users/luffy/Desktop/DL/data/hushen300/hushen300_data_deal_cut_test.xlsx'
    excelWriter=pd.ExcelWriter(excelPath,engine='openpyxl')
    excelAddSheet(temp2,excelWriter)







######created on 2018.7.28

#import pandas as pd
#import numpy as np
#from openpyxl import load_workbook
#excel_path = '/Users/luffy/Desktop/DL/data/hushen_standarize.xlsx'
#namelist=['open','high','low','close','volume','amt','pct_chg','turn']
#for j in range(len(namelist)):
#    temp = pd.read_excel(excel_path, sheetname=namelist[j])
#    temp2=temp[2792:]
#    def excelAddSheet(dataframe,excelWriter):
#        book = load_workbook(excelWriter.path)
#        excelWriter.book = book
#        dataframe.to_excel(excel_writer=excelWriter,sheet_name=namelist[j],index=None)
#        excelWriter.close()
#    excelPath='/Users/luffy/Desktop/DL/data/hushen_cut_test.xlsx'
#    excelWriter=pd.ExcelWriter(excelPath,engine='openpyxl')
#    excelAddSheet(temp2,excelWriter)
