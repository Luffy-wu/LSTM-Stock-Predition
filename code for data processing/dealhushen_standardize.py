#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Created on Sun Jul 29 11:48:04 2018

@author: luffy
"""
####revised on 2018.8.1

import pandas as pd
import numpy as np
from openpyxl import load_workbook
excel_path = '/Users/luffy/Desktop/DL/data/hushen300/hushen300_data_deal.xlsx'
namelist=['open','high','low','close','volume','amt','pct_chg','turn']
for i in range(len(namelist)):
    temp=pd.read_excel(excel_path, sheetname=namelist[i])
    temp2=pd.DataFrame(index=temp.index,columns=temp.columns)
    for j in range(len(temp.columns)):
        if j==0:
            temp2[temp.columns[j]]=temp[temp.columns[j]]
        if j>0:
            max_=max(temp[temp.columns[j]])
            min_=min(temp[temp.columns[j]])
            np_temp=np.array(temp[temp.columns[j]])
            np_temp=(np_temp-min_)/(max_-min_)
            temp2[temp.columns[j]]=np_temp
    def excelAddSheet(dataframe,excelWriter):
       book = load_workbook(excelWriter.path)
       excelWriter.book = book
       dataframe.to_excel(excel_writer=excelWriter,sheet_name=namelist[i],index=None)
       excelWriter.close()
    excelPath='/Users/luffy/Desktop/DL/data/hushen300/hushen300_data_deal_standarize.xlsx'
    excelWriter=pd.ExcelWriter(excelPath,engine='openpyxl')
    excelAddSheet(temp2,excelWriter)








######created on 2018.7.28
#open\high\low\close\volume\amt\pct_chg\turn
#import pandas as pd
#import numpy as np
#from openpyxl import load_workbook
#excel_path = '/Users/luffy/Desktop/DL/data/hushen.xlsx'
#temp = pd.read_excel(excel_path, sheetname='turn')
#temp2=pd.DataFrame(index=temp.index,columns=temp.columns)
#for j in range(len(temp.columns)):
#    if j==0:
#        temp2[temp.columns[j]]=temp[temp.columns[j]]
#    if j>0:
#        max_=max(temp[temp.columns[j]])
#        min_=min(temp[temp.columns[j]])
#        np_temp=np.array(temp[temp.columns[j]])
#        np_temp=(np_temp-min_)/(max_-min_)
#        temp2[temp.columns[j]]=np_temp
#def excelAddSheet(dataframe,excelWriter):
#   book = load_workbook(excelWriter.path)
#   excelWriter.book = book
#   dataframe.to_excel(excel_writer=excelWriter,sheet_name="turn",index=None)
#   excelWriter.close()
#excelPath='/Users/luffy/Desktop/DL/data/hushen_standarize.xlsx'
#excelWriter=pd.ExcelWriter(excelPath,engine='openpyxl')
#excelAddSheet(temp2,excelWriter)
##temp2.to_excel(writer,'open')
#



#############无用##########################33
#import pandas as pd
#excel_path = '/Users/luffy/Desktop/DL/data/hushen.xlsx'
#temp = pd.read_excel(excel_path, sheetname='open')
#temp2=pd.DataFrame(index=temp.index,columns=temp.columns)
#for j in range(len(temp.columns)):
#    if j==0:
#        temp2[temp.columns[j]]=temp[temp.columns[j]]
#    if j>0:
#        max_=max(temp[temp.columns[j]])
#        min_=min(temp[temp.columns[j]])
#        for i in range(len(temp)):
#            temp2[temp.columns[j]][i]=(temp[temp.columns[j]][i]-min_)/(max_-min_)
#
#
#
#writer = pd.ExcelWriter('/Users/luffy/Desktop/DL/data/hushen2.xlsx')
#temp2.to_excel(writer,'open')
#writer.save()